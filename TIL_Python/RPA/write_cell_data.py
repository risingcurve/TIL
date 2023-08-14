import openpyxl as op

wb = op.load_workbook(r"test.xlsx")
ws = wb["무"]

# "B1" cell에 입력하기
ws.cell(row=1, column=2).value = "입력테스트1"

# "C1" cell에 입력하기
ws["C1"].value = "입력테스트2"




# 임의의 숫자 리스트를 일렬로 세워 입력
# 임의의 숫자 리스트 정의
datalist = [2, 4, 8, 16, 32, 64, 128, 256]

# 행 값을 바꾸기 위한 인덱스 정의
i = 1

for data in datalist:
    # A열(column=1)에 행을 바꾸면서 입력
    ws.cell(row = i, column=1).value = data
    # 행값 증가
    i = i + 1

# 엑셀 파일 저장(파일명 : result.xlsx)
wb.save("result2.xlsx")
